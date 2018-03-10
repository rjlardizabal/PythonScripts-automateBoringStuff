
import sys
import openpyxl
from openpyxl.styles import Font


wb = openpyxl.Workbook()
sheet = wb.active


n = int(sys.argv[1])
boldFond = Font(bold=True)

for i in range(1, n+1):
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=1).font = boldFond

    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=1, column=i+1).font = boldFond


for i in range(1, n+1):
    for j in range(1, n+1):
        sheet.cell(row=i+1, column=j+1).value = i*j

wb.save('multiplicationTable.xlsx')
