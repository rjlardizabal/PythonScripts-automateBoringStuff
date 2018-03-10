import openpyxl
import pprint as pp


wb = openpyxl.load_workbook('multiplicationTable.xlsx')
sheet = wb.get_sheet_by_name('CellInverter')

rows = sheet.max_row
columns = sheet.max_column


table = [[sheet.cell(column=y, row=x).value for x in range(1, sheet.max_row+1)]
         for y in range(1, sheet.max_column+1)]
pp.pprint(table)

wb.create_sheet(title='Inverted')
sheetInv = wb.get_sheet_by_name('Inverted')

for rowNum in range(columns):
    for colNum in range(rows):
        print(f'Trying to acces {rowNum} {colNum} {table[rowNum][colNum]}')
        sheetInv.cell(row=rowNum+1,
                      column=colNum+1).value = table[rowNum][colNum]


wb.save('multiplicationTable.xlsx')
