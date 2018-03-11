#! python3
# sendDuesReminder.py - Sends emails based on payment status in srpeadsheet.

import openpyxl
import smtplib
import sys


wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}

for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=lastCol)
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('rj.lardizabal23@gmail.com', sys.argv[1])


for name, email in unpaidMembers.items():
    body = "Subject: {} dues unpaid.\n Dear {}, \n Records show that you have"
    "not paid dues for {month}. Please make this payment as soon as possible."
    "Thank you!".format(latestMonth, name, latestMonth)

    print('Sending email to {}...'.format(email))
    sendMailStatus = smtpObj.sendmail('rj.lardizabal23@gmail.com', email, body)

    if sendMailStatus != {}:
        print('There was a problem sending email to {}: {}'
              .format(email, sendMailStatus))
smtpObj.quit()
